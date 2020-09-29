import serial
import openpyxl
import time 

port = '/dev/ttyUSB0'  # define it to the port where the arduino is connected 
ser = serial.Serial(port = port, baudrate = 9600, timeout = 1)  #Serial port configuration
Workbook = openpyxl.Workbook()
WorkSheet = Workbook.active


 # number of inputs +1
#receive the data from the serial port and save it in the worksheet
num_line = 1
def get_data():
    global num_line
    
    num_column = 7
    ser.write(b'<')  # special character that the arduino should wait to star sending data
    
    #receive the data from the serial port, check if its a valid character and them save it in the right cell
    for i in range(1, num_column):
        serial_reading = ser.readline().decode('ascii')
        if serial_reading != '':
            WorkSheet.cell(row=num_line, column=i, value=serial_reading)
            print(f'leitura:{serial_reading} linha:{num_line}')
            if i == (num_column-1):
                num_line += 1
                print("somou")
    
    ser.write(b'>')  # special character that the arduino should wait to stop sending data
    
    Workbook.save("planilha.xlsx")


# checks if the outcome of the neural network is a valid possibility, and them send it to the arduino
def neural_network_outcome():
    result = WorkSheet.cell(num_line, 7)
    print(result)
    WorkSheet.cell(row = num_line, column = 7, value = 0)  #simula resposta da RN
    time.sleep(5)

#-----------------------------------------------------------------------------------------------------------------------
# main loop
while 1 :
    
    get_data()
    neural_network_outcome() 
    
