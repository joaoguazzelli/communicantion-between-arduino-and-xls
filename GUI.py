import tkinter as tk
import openpyxl
import serial


Workbook = openpyxl.Workbook()
WorkSheet = Workbook.active
root = tk.Tk()
root.geometry("680x400")
root.title('Eu tenho doutorado')
root.resizable(False, False)
root.config(padx=10, pady=10, background='#000')
port = '/dev/ttyACM0'  # define it to the port where the arduino is connected
ser = serial.Serial(port=port, baudrate=9600, timeout=1)  # Serial port configuration

variable1 = 0
variable2 = 0
variable3 = 0
variable4 = 0
variable5 = 0
variable6 = 0

sensors = [
    ['Fricção', variable1],
    ['Nutrição', variable2],
    ['Mobilidade', variable3],
    ['Atividade', variable4],
    ['Umidade', variable5],
    ['Percepção Sensorial', variable6]
]


def make_labels():
    sensors_row = []
    for column, column_value in enumerate(sensors):
        for row, row_value in enumerate(column_value):
            sensor = tk.Label(master=root, text=row_value)
            sensor.grid(row=row, column=column, sticky='news', padx=5, pady=5)
            if row == 1:
                sensor.configure(text=1, background='#f1f2f3')
            sensor.config(font=('Helvetica', 15, 'normal'), pady=40, background='#e1f2f3')
            sensors_row.append(sensor)
    
    return sensors_row

sensors_row = make_labels()
def get_data():
    num_line = 1
    num_column = 7
    ser.write(b'<')  # special character that the arduino should wait to star sending data

    # receive the data from the serial port, check if its a valid character and them save it on the right cell
    for i in range(1, num_column):
        serial_reading = ser.readline().decode('ascii')
        
        if serial_reading != '':
            WorkSheet.cell(row=num_line, column=i, value=serial_reading)
            reading = WorkSheet.cell(num_line, num_column)
            to_text = str(reading.value)
            sensors_row[1][num_column].configure(text=to_text)
            print(reading.value)
            
            if i == (num_column-1):
                num_line = num_line + 1

    ser.write(b'>')  # special character that the arduino should wait to stop sending data
    root.after(100, get_data)

get_data()
root.mainloop()
