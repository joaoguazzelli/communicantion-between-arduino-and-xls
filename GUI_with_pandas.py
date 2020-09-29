import tkinter as tk
import openpyxl
import serial
import time
import pandas as pd 
import xlrd 

Workbook = openpyxl.Workbook()
WorkSheet = Workbook.active
root = tk.Tk()
root.geometry("850x400")
root.title('Eu tenho Doutorado')
#root.resizable(False, False)
root.config(padx=10, pady=10, background='#000')
port = '/dev/ttyUSB0'  # define it to the port where the arduino is connected
ser = serial.Serial(port=port, baudrate=9600, timeout=1)  # Serial port configuration

variable1 = 0
variable2 = 0
variable3 = 0
variable4 = 0
variable5 = 0
variable6 = 0
variable7 = 0
sensors = [
    ['Fricção', variable1],
    ['Nutrição', variable2],
    ['Mobilidade', variable3],
    ['Atividade', variable4],
    ['Umidade', variable5],
    ['Percepção Sensorial', variable6],
    ['Resposta rede', variable7]
]


def make_labels():
    sensors_row = []
    for column, column_value in enumerate(sensors):
        for row, row_value in enumerate(column_value):
            sensor = tk.Label(master=root, text=row_value)
            sensor.grid(row=row, column=column, sticky='news', padx=5, pady=5)
            if row == 1:
                sensor.configure(text=1, background='#f1f2f3')
            sensor.config(font=('Helvetica', 15, 'normal'), pady=40, background='#e1f2f3')
            sensors_row.append(sensor)
    
    return sensors_row

def get_data():
    global num_line
    num_column = len(sensors) - 1
    
    contador_sensor = 1
    
    ser.write(b'<')  # special character that the arduino should wait to star sending data

    for i in range(1, num_column + 1): # receive the data from the serial port, check if its a valid character and them save it on the right cell
        serial_reading = ser.readline().decode('ascii')
        
        if serial_reading != '':
            WorkSheet.cell(row=num_line, column=i, value=serial_reading)
            sensors_row[contador_sensor].configure(text=serial_reading)
            
            if i == (num_column):
                num_line += 1
        
        contador_sensor += 2
    
    ser.write(b'>')  # special character that the arduino should wait to stop sending data
    Workbook.save("planilha.xlsx")


def neural_network():
    num_column = len(sensors)

    WorkSheet.cell(row=num_line, column=num_column, value=2)  # represents the neural network response

    network_response_text = WorkSheet.cell(row=num_line, column=num_column)
    
    if network_response_text.value == 1:
        sensors_row[-1].config(text='Risco Brando')
    
    elif network_response_text.value == 2:
        sensors_row[-1].config(text='Risco Moderado')
    
    elif network_response_text == 3:
        sensors_row[-1].config(text='Risco Severo')
    x = pd.read_excel("planilha.xlsx")
    print(x)


def GUI():
    get_data()
    neural_network()
    root.after(100, GUI)

num_line = 1

sensors_row = make_labels()
GUI()
root.mainloop()
