import serial
import openpyxl

porta = '/dev/ttyACM0'
ser = serial.Serial(port = porta, baudrate = 9600, timeout = 1)
wb = openpyxl.Workbook()
ws = wb.active

num_linha = 1
num_colunas = 7 # numero desejado de colunas antes da quebra de linhas +1

while 1 :
    ser.write(b'<') # caracter especial que deve ser aguardado pelo arduino para enviar os dado
    for i in range(1,num_colunas):
        leitura_serial = ser.readline().decode('ascii')
        print(f'leitura serial: {leitura_serial}')
        if leitura_serial != '':
            ws.cell(row = num_linha, column = i, value = leitura_serial)
            verifica = ws.cell(row=num_linha, column=i)
            print(f'leitura serial após gravação: {verifica.value}')
            if i == (num_colunas-1):
                num_linha = num_linha+1
                print(f'linha numero:{num_linha}')
                
    ser.write(b'>') # caracter especial que deve ser aguardado pelo arduino para enviar os dado
    wb.save("planilha.xlsx")